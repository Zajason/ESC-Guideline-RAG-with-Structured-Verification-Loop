from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

WORKSPACE_SITE_PACKAGES = Path(
    "/Users/zak/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/lib/python3.12/site-packages"
)
if WORKSPACE_SITE_PACKAGES.exists():
    sys.path.append(str(WORKSPACE_SITE_PACKAGES))

from openpyxl import load_workbook

from rag_core import (
    MODEL_NAME_DEFAULT,
    answer_query_A_single_pass,
    answer_query_A_with_verification,
    answer_query_B_single_pass,
    answer_query_B_with_verification,
)


DEFAULT_DOCX_DIR = Path("swisstransfer_ae864c5d-1a97-4925-85d5-d6a195c9eff7")
DEFAULT_EXCEL = Path("../Final AI & HFrEF patients list.xlsx")
DEFAULT_BATCH_MODEL = "gpt-5.4-mini"

RAG_COLUMNS = {
    # Query A
    "diag1": "DQ",
    "diag2": "DU",
    "diag3": "DY",
    "new_tni": "ED",
    "er_echo": "EI",
    "er_ct": "EN",
    "er_ctpa": "ES",
    "er_us": "EX",
    "discharge": "FC",
    "department_a": "FH",
    "prog_a1": "FL",
    "prog_a2": "FP",
    "prog_a3": "FT",
    # Query B
    "prog_b": "FY",
    "cause": "GD",
    "department_b": "GL",
    "hds": "GQ",
    "diuretics": "GZ",
    "diuretics_day": "HA",
    "diuretics_oral_day": "HG",
    "vasodilators": "HP",
    "vasodilators_day": "HQ",
    "intubation": "IA",
    "intubation_day": "IB",
    "niv": "IK",
    "niv_day": "IL",
    "abx": "IU",
    "abx_day": "IV",
    "inotropes": "JF",
    "inotropes_day": "JG",
    "ace_arb_arni": "JP",
    "ace_arb_arni_day": "JQ",
    "bb": "JZ",
    "bb_day": "KA",
    "sglt2": "KJ",
    "sglt2_day": "KK",
    "mra": "KT",
    "mra_day": "KU",
    "antiarrhythmic": "LD",
    "antiarrhythmic_day": "LE",
    "echo_b": "LN",
    "echo_b_day": "LO",
    "coro": "LX",
    "coro_day": "LY",
    "ctca": "ML",
    "ctca_day": "MM",
    "ct_b": "MV",
    "ct_b_day": "MW",
    "ctpa_b": "NF",
    "ctpa_b_day": "NG",
    "mri": "NP",
    "mri_in_hosp": "NQ",
    "icd": "NX",
    "icd_in_hosp": "NY",
    "interrogation": "OD",
    "holter": "ON",
    "holter_in_hosp": "OO",
    "aortic_valve": "OX",
    "aortic_method": "OY",
    "mitral_valve": "PH",
    "mitral_method": "PI",
    "major_error": "PM",
    "unsupported_data": "PQ",
    "major_error_text": "PU",
}

QUERY_A_EXCEL_CODE_MAP = {
    "new_tni_rag": "new_tni",
    "er_rag_echo": "er_echo",
    "er_rag_ct": "er_ct",
    "er_rag_ctpa": "er_ctpa",
    "er_rag_us": "er_us",
    "discharge_ai_rag": "discharge",
    "depar_rag": "department_a",
}

QUERY_B_EXCEL_CODE_MAP = {
    "prognosis_rag_final": "prog_b",
    "cause_decompensation_rag": "cause",
    "departm_rag": "department_b",
    "days_hosp_rag": "hds",
    "diuretics_rag": "diuretics",
    "diuretics_rag_day": "diuretics_day",
    "diuretics_rag_per_os_day": "diuretics_oral_day",
    "vasodilators_rag": "vasodilators",
    "vasodilators_rag_day": "vasodilators_day",
    "intubation_rag": "intubation",
    "intubation_rag_day": "intubation_day",
    "niv_rag": "niv",
    "niv_rag_day": "niv_day",
    "abx_rag": "abx",
    "abx_rag_day": "abx_day",
    "inotropes_rag": "inotropes",
    "inotropes_rag_day": "inotropes_day",
    "ace_arb_arni_rag": "ace_arb_arni",
    "ace_arb_arni_rag_day": "ace_arb_arni_day",
    "bb_rag": "bb",
    "bb_rag_day": "bb_day",
    "sglt2_rag": "sglt2",
    "sglt2_rag_day": "sglt2_day",
    "mra_rag": "mra",
    "mra_rag_day": "mra_day",
    "antiarrhythmic_rag": "antiarrhythmic",
    "antiarrhythmic_rag_day": "antiarrhythmic_day",
    "echo_rag": "echo_b",
    "echo_rag_day": "echo_b_day",
    "coro_rag": "coro",
    "coro_rag_day": "coro_day",
    "ctca_rag": "ctca",
    "ctca_rag_day": "ctca_day",
    "ct_rag": "ct_b",
    "ct_rag_day": "ct_b_day",
    "ctpa_rag": "ctpa_b",
    "ctpa_rag_day": "ctpa_b_day",
    "mri_rag": "mri",
    "mri_rag_in_hospital": "mri_in_hosp",
    "icd_rag": "icd",
    "icd_rag_in_hospital": "icd_in_hosp",
    "interrogation_rag": "interrogation",
    "holter_rag": "holter",
    "holter_rag_in_hospital": "holter_in_hosp",
    "aortic_valve_repair_rag": "aortic_valve",
    "aortic_method_repair": "aortic_method",
    "mitral_valve_repair_rag": "mitral_valve",
    "mitral_method_repair": "mitral_method",
}

ANSWER_SECTION_RE = re.compile(
    r"\n\s*(Answers?\s+(CHATGPT|CHAT|DEEP|C\+|RAG)|CHATGPT\s+ANSWERS?|DEEP\s+ANSWERS?|AI\s*C\+\s*ANSWERS?)\s*:",
    flags=re.I,
)
POST_DX_RE = re.compile(r"\bNow I will tell you\b", flags=re.I)


@dataclass
class PatientDoc:
    path: Path
    code_number: int
    code_text: str
    initials: str
    query_a_text: str
    query_b_addon: str


def docx_text(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    parts: list[str] = []
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            parts.append(line)
    return "\n".join(parts).strip()


def parse_patient_doc(path: Path) -> PatientDoc:
    raw = docx_text(path)
    raw = ANSWER_SECTION_RE.split(raw, maxsplit=1)[0].strip()
    post_match = POST_DX_RE.search(raw)
    if not post_match:
        raise ValueError("Could not find post-diagnosis marker: 'Now I will tell you'")

    query_a = raw[: post_match.start()].strip()
    query_b = raw[post_match.start() :].strip()

    filename_match = re.match(r"\s*(\d+)", path.stem)
    code_number = int(filename_match.group(1)) if filename_match else 0
    tail = re.sub(r"^\s*\d+\s*[-.]\s*", "", path.stem).strip()
    initials = normalize_initials(tail)

    code_match = re.search(r"CODE\s*:\s*([^\n]+)", raw, flags=re.I)
    code_text = code_match.group(1).strip() if code_match else path.stem

    return PatientDoc(
        path=path,
        code_number=code_number,
        code_text=code_text,
        initials=initials,
        query_a_text=query_a,
        query_b_addon=query_b,
    )


def normalize_initials(text: str) -> str:
    text = unicodedata.normalize("NFD", text.upper())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^A-ZΑ-Ω]", "", text)


def row_initials(ws, row: int) -> str:
    surname = str(ws[f"C{row}"].value or "").strip()
    name = str(ws[f"D{row}"].value or "").strip()
    return normalize_initials((surname[:1] or "") + (name[:1] or ""))


def all_text(obj: Any) -> str:
    chunks: list[str] = []

    def walk(x: Any) -> None:
        if isinstance(x, str):
            chunks.append(x)
        elif isinstance(x, dict):
            for v in x.values():
                walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)

    walk(obj)
    return "\n".join(chunks).lower()


def get_score(item: dict[str, Any]) -> int | None:
    if not isinstance(item, dict):
        return None
    value = item.get("score_1_to_8", item.get("score_1_to_7"))
    try:
        return int(value)
    except Exception:
        return None


def day_from_text(text: str) -> int | None:
    m = re.search(r"\bday\s*([0-9]+)\b", text, flags=re.I)
    if m:
        return int(m.group(1))
    if re.search(r"\bER\b|\badmission\b|\bimmediate\b|\bnow\b", text, flags=re.I):
        return 0
    if re.search(r"\bdischarge\b|\boutpatient\b", text, flags=re.I):
        return None
    return None


def first_day_for_terms(blocks: list[dict[str, Any]], terms: list[str]) -> int | None:
    for block in blocks:
        text = all_text(block)
        if any(term in text for term in terms):
            day = day_from_text(text)
            if day is not None:
                return day
    return None


def contains_any(text: str, terms: list[str]) -> bool:
    return any(term in text for term in terms)


def as_int_or_none(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except Exception:
        return None


def copy_excel_codes(excel_codes: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    if not isinstance(excel_codes, dict):
        return out
    for source_key, target_key in mapping.items():
        if source_key in excel_codes:
            out[target_key] = as_int_or_none(excel_codes.get(source_key))
    return out


def clamp_int(value: Any, allowed: set[int], default: int | None = None) -> int | None:
    ivalue = as_int_or_none(value)
    if ivalue in allowed:
        return ivalue
    return default


def action_text(block: dict[str, Any]) -> str:
    if not isinstance(block, dict):
        return ""
    parts: list[str] = []
    for key in ["day", "actions", "action", "item", "purpose", "clinical_judgment"]:
        value = block.get(key)
        if isinstance(value, list):
            parts.extend(str(v) for v in value)
        elif value is not None:
            parts.append(str(value))
    return " ".join(parts).lower()


def is_conditional_only(text: str) -> bool:
    conditional_terms = [
        "if worsening", "if deterior", "if hypoxaemia develops", "if hypoxemia develops",
        "if respiratory failure", "if indicated", "if needed", "if necessary",
        "if respiratory distress", "if progressive", "if persistent",
        "only if", "may be considered", "can be considered", "consider if",
        "low threshold", "escalate if", "unless", "possible need",
    ]
    return any(term in text for term in conditional_terms)


def blocks_with_terms(blocks: list[dict[str, Any]], terms: list[str]) -> list[dict[str, Any]]:
    return [block for block in blocks if contains_any(action_text(block), terms)]


def definite_blocks_with_terms(blocks: list[dict[str, Any]], terms: list[str]) -> list[dict[str, Any]]:
    return [block for block in blocks_with_terms(blocks, terms) if not is_conditional_only(action_text(block))]


def any_definite_action(blocks: list[dict[str, Any]], terms: list[str]) -> bool:
    return bool(definite_blocks_with_terms(blocks, terms))


def diagnosis_code(dx: str) -> int:
    t = dx.lower()
    if "pulmonary embol" in t or re.search(r"\bpe\b", t):
        return 8
    if "stemi" in t:
        return 5
    if "nstemi" in t or "nste" in t or "acute coronary" in t or "acs" in t:
        return 6
    if "respiratory infection" in t or "pneumonia" in t:
        return 11 if ("heart failure" in t or "hf" in t) else 10
    if "copd" in t:
        return 9
    if "aortic" in t or "mitral" in t or "valve" in t or "valvular" in t:
        return 12
    if "ventricular tach" in t or "nsvt" in t or re.search(r"\bvt\b", t):
        return 13
    if "atrial fibrillation" in t or re.search(r"\baf\b", t):
        return 14
    if "hfmref" in t or "mid-range" in t or "mildly reduced" in t:
        return 2
    if "hfpef" in t or "preserved" in t:
        return 3
    if "hfr" in t or "reduced ejection" in t:
        return 1
    if "heart failure" in t or "adhf" in t or "pulmonary oedema" in t or "pulmonary edema" in t:
        return 4
    return 15


def department_code_a(value: str) -> int | None:
    t = value.upper()
    if "CICU" in t:
        return 2
    if "CARDIOLOGY" in t:
        return 1
    if "RESP" in t or "PULMON" in t:
        return 3
    if "INTERNAL" in t:
        return 4
    if "ICU" in t:
        return 5
    if "SURG" in t:
        return 6
    return None


def department_code_b(value: str) -> int | None:
    t = value.upper()
    if "CICU" in t:
        return 1
    if "CARDIOLOGY" in t:
        return 0
    return None


def disposition_discharge_code(answer: str) -> int:
    return 1 if "DISCHARGE" in answer.upper() else 0


def cause_code(text: str) -> int:
    t = text.lower()
    if "no decomp" in t or "not decomp" in t:
        return 9
    if "end stage" in t or "inotrope dependent" in t:
        return 11
    if "valve" in t or "aortic stenosis" in t or "mitral" in t:
        return 12
    if "hypertensive" in t or "hypertension" in t or "blood pressure" in t or "afterload" in t:
        return 8
    if "type 2" in t or "minoca" in t or "demand ischem" in t or "myocardial injury" in t:
        return 7
    if "drug" in t or "medication" in t or "nonadherence" in t or "nsaid" in t:
        return 6
    if "diet" in t or "salt" in t or "fluid" in t:
        return 5
    if "infection" in t or "pneumonia" in t or "sepsis" in t:
        return 4
    if "tachycardia" in t or "rapid ventricular" in t:
        return 3
    if "arrhythm" in t or "atrial fibrillation" in t or re.search(r"\baf\b", t):
        return 2
    if "volume overload" in t or "oedema" in t or "edema" in t or "congestion" in t:
        return 1
    return 10


def er_us_code(text: str) -> int:
    if "venous" in text or "dvt" in text or "limb" in text:
        return 4
    if "abdomen" in text or "abdominal" in text:
        return 2
    if "brain" in text or "cranial" in text:
        return 3
    return 1


def query_a_values(resp: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    excel_codes = resp.get("excel_codes", {})
    if isinstance(excel_codes, dict):
        diagnoses = excel_codes.get("diagnoses")
        if isinstance(diagnoses, list):
            for i, key in enumerate(["diag1", "diag2", "diag3"]):
                if i < len(diagnoses):
                    out[key] = as_int_or_none(diagnoses[i])
        prognosis = excel_codes.get("prognosis_rag")
        if isinstance(prognosis, list):
            for i, key in enumerate(["prog_a1", "prog_a2", "prog_a3"]):
                if i < len(prognosis):
                    out[key] = as_int_or_none(prognosis[i])
        out.update(copy_excel_codes(excel_codes, QUERY_A_EXCEL_CODE_MAP))

    differential = resp.get("differential", [])
    for i, key in enumerate(["diag1", "diag2", "diag3"]):
        if key not in out and i < len(differential):
            out[key] = diagnosis_code(str(differential[i].get("dx", "")))

    tests_text = all_text(resp.get("tests", []))
    out.setdefault("new_tni", 1 if contains_any(tests_text, ["troponin", "tni"]) else 0)
    out.setdefault("er_echo", 1 if contains_any(tests_text, ["echo", "echocardi", "tte"]) else 0)
    out.setdefault("er_ctpa", 1 if contains_any(tests_text, ["ctpa", "pulmonary angi"]) else 0)
    out.setdefault("er_ct", 1 if contains_any(tests_text, ["ct thorax", "chest ct", "ct chest", "thoracic ct"]) else 0)
    out.setdefault("er_us", er_us_code(tests_text) if contains_any(tests_text, ["ultrasound", "doppler", "lung us"]) else 0)

    disposition = resp.get("disposition", {})
    if isinstance(disposition, dict):
        out.setdefault("discharge", disposition_discharge_code(str(disposition.get("answer", ""))))
        out.setdefault("department_a", department_code_a(str(disposition.get("department", ""))))

    prognosis = resp.get("prognosis", [])
    for i, key in enumerate(["prog_a1", "prog_a2", "prog_a3"]):
        if i < len(prognosis):
            out[key] = get_score(prognosis[i])
    normalize_query_a_codebook(out)
    return out


def normalize_query_a_codebook(out: dict[str, Any]) -> None:
    allowed_sets = {
        "diag1": set(range(1, 16)),
        "diag2": set(range(1, 16)),
        "diag3": set(range(1, 16)),
        "er_ct": {0, 1, 2, 3},
        "er_us": {0, 1, 2, 3, 4},
        "department_a": {1, 2, 3, 4, 5, 6},
        "prog_a1": set(range(1, 9)),
        "prog_a2": set(range(1, 9)),
        "prog_a3": set(range(1, 9)),
    }
    for key in ["new_tni", "er_echo", "er_ctpa", "discharge"]:
        if key in out:
            out[key] = 1 if as_int_or_none(out[key]) and as_int_or_none(out[key]) > 0 else 0
    for key, allowed in allowed_sets.items():
        if key in out:
            out[key] = clamp_int(out.get(key), allowed)


def treatment_blocks(resp: dict[str, Any]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for key in ["cause_workup_algorithm", "treatment_plan_by_day", "safety_critical"]:
        value = resp.get(key, [])
        if isinstance(value, list):
            blocks.extend([x for x in value if isinstance(x, dict)])
    followup = resp.get("followup", {})
    if isinstance(followup, dict):
        for key in ["in_hospital", "after_discharge"]:
            value = followup.get(key, [])
            if isinstance(value, list):
                blocks.extend([x for x in value if isinstance(x, dict)])
    return blocks


def query_b_values(resp: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    excel_codes = resp.get("excel_codes", {})
    if isinstance(excel_codes, dict):
        out.update(copy_excel_codes(excel_codes, QUERY_B_EXCEL_CODE_MAP))

    out.setdefault("prog_b", get_score(resp.get("prognosis", {})))
    trigger = resp.get("most_likely_trigger", {})
    if isinstance(trigger, dict):
        out.setdefault("cause", cause_code(str(trigger.get("answer", ""))))
    admission = resp.get("admission_level", {})
    if isinstance(admission, dict):
        out.setdefault("department_b", department_code_b(str(admission.get("department", ""))))
    try:
        estimated_duration = resp.get("estimated_duration", {})
        if isinstance(estimated_duration, dict):
            out.setdefault("hds", int(estimated_duration.get("hds_score_1_to_5")))
    except Exception:
        out.setdefault("hds", None)

    blocks = treatment_blocks(resp)
    text = " ".join(action_text(block) for block in blocks)

    diuretic_terms = ["diuretic", "furosemide", "torasemide", "bumetanide"]
    if "diuretics" not in out:
        definite_diuretic = definite_blocks_with_terms(blocks, diuretic_terms)
        if definite_diuretic:
            definite_text = " ".join(action_text(block) for block in definite_diuretic)
            out["diuretics"] = 1 if contains_any(definite_text, ["iv ", "i.v", "intravenous"]) else 2
            out["diuretics_day"] = first_day_for_terms(definite_diuretic, diuretic_terms)
        else:
            out["diuretics"] = 0

    if "diuretics_oral_day" not in out and contains_any(text, ["oral diuretic", "per os diuretic", "switch to oral", "transition to oral"]):
        out["diuretics_oral_day"] = first_day_for_terms(blocks, ["oral diuretic", "switch to oral", "transition to oral"])

    simple_terms = {
        "vasodilators": ["vasodilator", "nitrate", "nitroglycerin", "nitroprusside"],
        "niv": ["non-invasive ventilation", "non invasive ventilation", "niv", "cpap", "bipap"],
        "abx": ["antibiotic", "antimicrobial", "azithromycin", "ceftriaxone"],
        "antiarrhythmic": ["antiarrhythmic", "amiodarone", "cardioversion"],
        "echo_b": ["echo", "echocardi", "tte", "toe", "transesophageal"],
        "ctpa_b": ["ctpa", "pulmonary angi"],
        "ctca": ["ctca", "coronary ct", "ct coronary"],
        "ct_b": ["ct thorax", "chest ct", "ct chest", "ct abdomen", "brain ct"],
        "mri": ["cardiac mri", "cmr", "magnetic resonance"],
        "interrogation": ["interrogation", "device check", "icd check", "crt-d check", "crtd check"],
        "holter": ["holter", "ambulatory ecg"],
    }
    for key, terms in simple_terms.items():
        if key not in out:
            out[key] = 1 if any_definite_action(blocks, terms) else 0
        day_key = f"{key}_day"
        if key in ["mri", "holter"]:
            out.setdefault(f"{key}_in_hosp", out[key])
        elif day_key in RAG_COLUMNS and out[key]:
            out.setdefault(day_key, first_day_for_terms(definite_blocks_with_terms(blocks, terms), terms))

    if "intubation" not in out:
        intubation_blocks = blocks_with_terms(blocks, ["intubat", "mechanical ventilation"])
        if intubation_blocks:
            definite_intubation = [b for b in intubation_blocks if not is_conditional_only(action_text(b))]
            out["intubation"] = 1 if definite_intubation else 2
            out["intubation_day"] = first_day_for_terms(definite_intubation or intubation_blocks, ["intubat", "mechanical ventilation"])
        else:
            out["intubation"] = 0

    if "inotropes" not in out:
        inotrope_blocks = definite_blocks_with_terms(blocks, ["inotrope", "dobutamine", "milrinone", "vasopressor", "noradrenaline", "norepinephrine"])
        if inotrope_blocks:
            inotrope_text = " ".join(action_text(block) for block in inotrope_blocks)
            low_bp = contains_any(inotrope_text, ["low bp", "hypotension", "hypoperfusion", "shock"])
            diuresis_support = contains_any(inotrope_text, ["support diuresis", "facilitate diuresis", "during diuresis"])
            out["inotropes"] = 4 if low_bp and diuresis_support else (2 if low_bp else 1)
            out["inotropes_day"] = first_day_for_terms(inotrope_blocks, ["inotrope", "dobutamine", "milrinone", "vasopressor"])
        else:
            out["inotropes"] = 0

    gdmt_terms = {
        "ace_arb_arni": ["ace inhibitor", "acei", "arb", "arni", "sacubitril", "valsartan"],
        "bb": ["beta-blocker", "beta blocker", "bisoprolol", "carvedilol", "metoprolol", "nebivolol"],
        "sglt2": ["sglt2", "sglt-2", "dapagliflozin", "empagliflozin"],
        "mra": ["mra", "mineralocorticoid", "spironolactone", "eplerenone"],
    }
    for key, terms in gdmt_terms.items():
        if key not in out:
            out[key] = 1 if any_definite_action(blocks, terms) else 0
        if out[key]:
            out.setdefault(f"{key}_day", first_day_for_terms(definite_blocks_with_terms(blocks, terms), terms))

    if "coro" not in out:
        coro_terms = ["coronary angiography", "invasive coronary", "coronarography"]
        coro_blocks = definite_blocks_with_terms(blocks, coro_terms)
        out["coro"] = 1 if coro_blocks else 0
        if coro_blocks:
            out["coro_day"] = first_day_for_terms(coro_blocks, coro_terms)

    if "icd" not in out and any_definite_action(blocks, ["icd", "crt", "defibrillator"]):
        if contains_any(text, ["already icd", "existing icd", "has icd", "crt-d", "crtd"]):
            out["icd"] = 5
        elif contains_any(text, ["upgrade"]):
            out["icd"] = 6
        elif contains_any(text, ["3 months", "three months"]):
            out["icd"] = 3
        elif contains_any(text, ["40 days", "forty days"]):
            out["icd"] = 2
        else:
            out["icd"] = 1
        out["icd_in_hosp"] = 1 if contains_any(text, ["in hospital", "during hospitalization"]) else 0
    else:
        out.setdefault("icd", 0)

    if "aortic_valve" not in out and any_definite_action(blocks, ["aortic valve", "tavi", "savr"]):
        out["aortic_valve"] = 3 if contains_any(text, ["heart team", "frailty"]) else 1
        out["aortic_method"] = 2 if contains_any(text, ["savr", "surgical aortic"]) else 1
    else:
        out.setdefault("aortic_valve", 0)

    if "mitral_valve" not in out and any_definite_action(blocks, ["mitral valve", "m-teer", "mteer", "mitraclip", "surgical mitral"]):
        out["mitral_valve"] = 3 if contains_any(text, ["heart team", "frailty"]) else 1
        out["mitral_method"] = 2 if contains_any(text, ["surgical mitral", "surgery"]) else 1
    else:
        out.setdefault("mitral_valve", 0)

    normalize_query_b_codebook(out, blocks)
    rescue_query_b_prognosis_from_response(out, resp)
    return out


def rescue_query_b_prognosis_from_response(out: dict[str, Any], resp: dict[str, Any]) -> None:
    text = all_text(resp)
    current = as_int_or_none(out.get("prog_b")) or 0
    hds = as_int_or_none(out.get("hds")) or 0
    cause = as_int_or_none(out.get("cause"))
    rescued = current

    if re.search(r"\blactate\s*(?:=|of)?\s*[67](?:\.\d+)?", text) or "very high lactate" in text:
        rescued = max(rescued, 8)
    if ("gcs 8" in text or "gcs=8" in text) and contains_any(text, ["severe renal", "creatinine", "azotemia", "aki", "ckd"]):
        rescued = max(rescued, 8)
    if cause == 11 and hds == 5 and contains_any(text, ["end-stage", "inotrope-dependent"]):
        rescued = max(rescued, 8)
    if contains_any(text, ["high-grade av block", "3:1 conduction", "ventricular rate of 30", "ventricular rate 30"]):
        rescued = max(rescued, 8)
    if rescued != current:
        out["prog_b"] = rescued


def normalize_query_b_codebook(out: dict[str, Any], blocks: list[dict[str, Any]]) -> None:
    """Keep model-supplied excel_codes inside the workbook's expected codebooks."""
    simple_binary_fields = [
        "abx", "ace_arb_arni", "bb", "sglt2", "mra", "antiarrhythmic",
        "echo_b", "ctca", "ct_b", "mri", "mri_in_hosp", "icd_in_hosp",
        "interrogation", "holter", "holter_in_hosp",
    ]
    for key in simple_binary_fields:
        if key in out:
            out[key] = 1 if as_int_or_none(out[key]) and as_int_or_none(out[key]) > 0 else 0

    vasodilator_terms = ["vasodilator", "nitrate", "nitroglycerin", "nitroprusside"]
    if clamp_int(out.get("vasodilators"), {0, 1}) is None:
        out["vasodilators"] = 1 if blocks_with_terms(blocks, vasodilator_terms) else 0

    niv_terms = ["non-invasive ventilation", "non invasive ventilation", "niv", "cpap", "bipap"]
    if clamp_int(out.get("niv"), {0, 1}) is None:
        out["niv"] = 1 if any_definite_action(blocks, niv_terms) else 0

    if not out.get("vasodilators"):
        out.pop("vasodilators_day", None)
    if not out.get("niv"):
        out.pop("niv_day", None)

    allowed_sets = {
        "prog_b": set(range(1, 9)),
        "cause": set(range(1, 13)),
        "department_b": {0, 1},
        "hds": set(range(1, 7)),
        "diuretics": {0, 1, 2},
        "intubation": {0, 1, 2},
        "ctpa_b": {0, 1, 2},
        "inotropes": {0, 1, 2, 3, 4},
        "coro": {0, 1},
        "icd": {0, 1, 2, 3, 4, 5, 6},
        "aortic_valve": {0, 1, 2, 3},
        "aortic_method": {1, 2},
        "mitral_valve": {0, 1, 2, 3},
        "mitral_method": {1, 2},
    }
    defaults = {
        "prog_b": 4,
        "cause": 10,
        "department_b": 0,
        "hds": 3,
        "diuretics": 0,
        "intubation": 0,
        "ctpa_b": 0,
        "inotropes": 0,
        "coro": 1 if any_definite_action(blocks, ["coronary angiography", "invasive coronary", "coronarography"]) else 0,
        "icd": 0,
        "aortic_valve": 0,
        "mitral_valve": 0,
    }
    for key, allowed in allowed_sets.items():
        if key in out:
            out[key] = clamp_int(out.get(key), allowed, defaults.get(key))

    for key in ["aortic_method", "mitral_method"]:
        if key in out and out[key] is None:
            out.pop(key, None)


def compact_debug(debug: dict[str, Any]) -> dict[str, Any]:
    return {
        "base_queries": debug.get("base_queries", debug.get("queries")),
        "final_queries": debug.get("final_queries"),
        "case_snapshot": debug.get("case_snapshot"),
        "verifier_reports": debug.get("verifier_reports"),
    }


def find_excel_rows(ws) -> dict[int, int]:
    rows: dict[int, int] = {}
    for row in range(3, ws.max_row + 1):
        value = ws[f"B{row}"].value
        if isinstance(value, (int, float)):
            rows[int(value)] = row
    return rows


def find_excel_initial_rows(ws) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    for row in range(3, ws.max_row + 1):
        initials = row_initials(ws, row)
        if initials:
            rows.setdefault(initials, []).append(row)
    return rows


def find_excel_row_for_doc(ws, row_by_code: dict[int, int], row_by_initials: dict[str, list[int]], doc: PatientDoc) -> int | None:
    row = row_by_code.get(doc.code_number)
    if row is not None and (not doc.initials or row_initials(ws, row) == doc.initials):
        return row

    matches = row_by_initials.get(doc.initials, [])
    if len(matches) == 1:
        return matches[0]
    return row


def write_values(ws, row: int, values: dict[str, Any], overwrite: bool) -> None:
    for key, value in values.items():
        if key not in RAG_COLUMNS or value is None:
            continue
        cell = ws[f"{RAG_COLUMNS[key]}{row}"]
        if overwrite or cell.value in (None, ""):
            cell.value = value


def rag_row_complete(ws, row: int) -> bool:
    required = ["DQ", "DU", "DY", "FC", "FH", "FL", "FP", "FT", "FY", "GD", "GQ"]
    return all(ws[f"{col}{row}"].value not in (None, "") for col in required)


def run_one(doc: PatientDoc, model: str, verified: bool, repair_on_format: bool = True) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    if verified:
        resp_a, debug_a = answer_query_A_with_verification(doc.query_a_text, model=model)
        resp_b, debug_b = answer_query_B_with_verification(
            doc.query_a_text + "\n\n" + doc.query_b_addon,
            model=model,
        )
    else:
        try:
            resp_a, debug_a = answer_query_A_single_pass(doc.query_a_text, model=model)
            resp_b, debug_b = answer_query_B_single_pass(
                doc.query_a_text + "\n\n" + doc.query_b_addon,
                model=model,
            )
        except ValueError:
            if not repair_on_format:
                raise
            resp_a, debug_a = answer_query_A_with_verification(doc.query_a_text, model=model, max_rounds=1)
            resp_b, debug_b = answer_query_B_with_verification(
                doc.query_a_text + "\n\n" + doc.query_b_addon,
                model=model,
                max_rounds=1,
            )
    values = {}
    values.update(query_a_values(resp_a))
    values.update(query_b_values(resp_b))
    return values, {"response": resp_a, "debug": compact_debug(debug_a)}, {"response": resp_b, "debug": compact_debug(debug_b)}


def is_quota_error(exc: Exception) -> bool:
    text = repr(exc).lower()
    return "insufficient_quota" in text or "exceeded your current quota" in text


def is_auth_error(exc: Exception) -> bool:
    text = repr(exc).lower()
    return "authenticationerror" in text or "invalid_api_key" in text


def is_retryable_error(exc: Exception) -> bool:
    text = repr(exc).lower()
    retry_terms = [
        "ratelimiterror",
        "api connection",
        "apiconnectionerror",
        "timeout",
        "temporarily unavailable",
        "server error",
        "service unavailable",
        "429",
        "500",
        "502",
        "503",
        "504",
    ]
    return not is_quota_error(exc) and not is_auth_error(exc) and any(term in text for term in retry_terms)


def run_one_with_retries(
    doc: PatientDoc,
    model: str,
    verified: bool,
    max_retries: int,
    retry_sleep: float,
    repair_on_format: bool,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    attempt = 0
    while True:
        try:
            return run_one(doc, model=model, verified=verified, repair_on_format=repair_on_format)
        except Exception as exc:
            if attempt >= max_retries or not is_retryable_error(exc):
                raise
            delay = retry_sleep * (2 ** attempt)
            print(f"  Retryable error on attempt {attempt + 1}: {exc!r}. Sleeping {delay:.1f}s.", flush=True)
            time.sleep(delay)
            attempt += 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Run ESC RAG Query A/B for DOCX patients and fill only RAG columns in the Excel workbook.")
    parser.add_argument("--docx-dir", type=Path, default=DEFAULT_DOCX_DIR)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--model", default=None, help=f"Defaults to ESC_RAG_MODEL or {DEFAULT_BATCH_MODEL}.")
    parser.add_argument("--single-pass", action="store_true", help="Deprecated/no-op: single-pass is now the cost-conscious default.")
    parser.add_argument("--verified", action="store_true", help="Use the verifier/reviser loop. More accurate for citations/format, but costs more API calls.")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--only", nargs="*", type=int, default=None, help="Run only these numeric patient codes.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing RAG cells. Default fills blanks only.")
    parser.add_argument("--rerun-complete", action="store_true", help="Call the API even when the required RAG columns already look complete.")
    parser.add_argument("--max-retries", type=int, default=3, help="Retries for transient API/network/rate-limit errors.")
    parser.add_argument("--retry-sleep", type=float, default=20.0, help="Base seconds for exponential backoff.")
    parser.add_argument("--no-format-repair", action="store_true", help="Do not use a one-round verified repair when single-pass output has an invalid shape.")
    parser.add_argument("--dry-run", action="store_true", help="Run extraction/mapping without calling RAG or writing Excel.")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    docx_dir = (root / args.docx_dir).resolve() if not args.docx_dir.is_absolute() else args.docx_dir
    excel_path = (root / args.excel).resolve() if not args.excel.is_absolute() else args.excel
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    output_path = args.output or excel_path.with_name(excel_path.stem + f"_RAG_filled_{timestamp}" + excel_path.suffix)
    output_path = output_path.resolve()
    log_dir = root / "batch_runs" / timestamp
    log_dir.mkdir(parents=True, exist_ok=True)

    docs = [parse_patient_doc(p) for p in sorted(docx_dir.glob("*.docx"), key=lambda p: int(re.match(r"\s*(\d+)", p.stem).group(1)))]
    if args.only:
        allowed = set(args.only)
        docs = [d for d in docs if d.code_number in allowed]
    if args.limit:
        docs = docs[: args.limit]

    if args.dry_run:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        row_by_code = find_excel_rows(ws)
        row_by_initials = find_excel_initial_rows(ws)
        for doc in docs:
            row = find_excel_row_for_doc(ws, row_by_code, row_by_initials, doc)
            row_note = f"row={row}" if row else "row=NOT FOUND"
            print(f"{doc.code_number}: {doc.path.name} | initials={doc.initials} | {row_note} | A chars={len(doc.query_a_text)} | B addon chars={len(doc.query_b_addon)}")
        return 0

    if output_path.exists() and output_path.resolve() == excel_path.resolve():
        pass
    else:
        shutil.copy2(excel_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    row_by_code = find_excel_rows(ws)
    row_by_initials = find_excel_initial_rows(ws)
    failures: list[dict[str, Any]] = []
    model = args.model or os.environ.get("ESC_RAG_MODEL") or DEFAULT_BATCH_MODEL or MODEL_NAME_DEFAULT
    skipped: list[dict[str, Any]] = []

    for idx, doc in enumerate(docs, start=1):
        print(f"[{idx}/{len(docs)}] Running patient {doc.code_text} from {doc.path.name}", flush=True)
        row = find_excel_row_for_doc(ws, row_by_code, row_by_initials, doc)
        if row is None:
            failures.append({"code": doc.code_text, "file": str(doc.path), "error": "No matching Excel row in column B"})
            continue
        if not args.overwrite and not args.rerun_complete and rag_row_complete(ws, row):
            skipped.append({"code": doc.code_text, "file": str(doc.path), "excel_row": row, "reason": "RAG row already complete"})
            print("  Skipping: RAG columns already complete.", flush=True)
            continue
        try:
            values, raw_a, raw_b = run_one_with_retries(
                doc,
                model=model,
                verified=args.verified and not args.single_pass,
                max_retries=args.max_retries,
                retry_sleep=args.retry_sleep,
                repair_on_format=not args.no_format_repair,
            )
            write_values(ws, row, values, overwrite=args.overwrite)
            wb.save(output_path)
            patient_log = {
                "code": doc.code_text,
                "file": str(doc.path),
                "excel_row": row,
                "mapped_values": values,
                "query_a": raw_a,
                "query_b": raw_b,
            }
            log_name = re.sub(r"[^0-9A-Za-zΑ-Ωα-ω]+", "_", doc.path.stem)
            (log_dir / f"{log_name}_rag.json").write_text(json.dumps(patient_log, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            failures.append({"code": doc.code_text, "file": str(doc.path), "error": repr(exc)})
            log_name = re.sub(r"[^0-9A-Za-zΑ-Ωα-ω]+", "_", doc.path.stem)
            (log_dir / f"{log_name}_ERROR.json").write_text(
                json.dumps(failures[-1], ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"  ERROR: {exc!r}", file=sys.stderr, flush=True)
            if is_auth_error(exc):
                print("Fatal OpenAI authentication error; stopping batch run.", file=sys.stderr, flush=True)
                break
            if is_quota_error(exc):
                print("Fatal OpenAI quota error; stopping batch run so it can be resumed later.", file=sys.stderr, flush=True)
                break

    wb.save(output_path)
    summary = {
        "input_excel": str(excel_path),
        "output_excel": str(output_path),
        "docx_dir": str(docx_dir),
        "patients_requested": len(docs),
        "patients_skipped_complete": len(skipped),
        "skipped": skipped,
        "failures": failures,
        "mode": "verified" if args.verified and not args.single_pass else "single_pass",
        "model": model,
    }
    (log_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
