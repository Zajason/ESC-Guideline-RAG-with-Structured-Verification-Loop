#!/usr/bin/env python3
"""Evaluate final prognosis columns against the doctor/result prognosis.

Example:
    python scripts/evaluate_models.py "../Final AI & HFrEF patients list_RAG_refilled_gpt54mini_rerun.xlsx"
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - user setup guard
    raise SystemExit("Missing dependency: openpyxl. Install with `pip install -r requirements.txt`.") from exc

from esc_rag.metrics import confusion_matrix, per_class_recall, prognosis_metrics


DEFAULT_COLUMNS = {
    "doctor": "GE",
    "CHAT": "FV",
    "DEEP": "FW",
    "C+": "FX",
    "RAG": "FY",
}


def cell_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        return None


def load_pairs(path: Path, columns: dict[str, str], start_row: int, end_row: int) -> dict[str, list[tuple[int, int]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    doctor_col = columns["doctor"]
    out: dict[str, list[tuple[int, int]]] = {}
    for model, pred_col in columns.items():
        if model == "doctor":
            continue
        pairs: list[tuple[int, int]] = []
        for row in range(start_row, end_row + 1):
            actual = cell_int(ws[f"{doctor_col}{row}"].value)
            pred = cell_int(ws[f"{pred_col}{row}"].value)
            if actual is not None and pred is not None:
                pairs.append((pred, actual))
        out[model] = pairs
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--start-row", type=int, default=3)
    parser.add_argument("--end-row", type=int, default=79)
    parser.add_argument("--json", action="store_true", help="Emit machine-readable JSON.")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    pairs_by_model = load_pairs(args.workbook, DEFAULT_COLUMNS, args.start_row, args.end_row)
    report = {}
    for model, pairs in pairs_by_model.items():
        report[model] = {
            "summary": prognosis_metrics(pairs),
            "per_class_recall": per_class_recall(pairs),
            "confusion_matrix": confusion_matrix(pairs),
        }

    if args.json:
        print(json.dumps(report, indent=2, sort_keys=True))
        return 0

    print(f"Workbook: {args.workbook}")
    print("Compared model final prognosis columns against doctor/result column GE.\n")
    print(f"{'Model':<8} {'N':>3} {'Exact':>8} {'MAE':>7} {'RMSE':>7} {'Within1':>8} {'Bias':>7}")
    for model, data in sorted(report.items(), key=lambda item: item[1]["summary"]["exact_accuracy"], reverse=True):
        summary = data["summary"]
        print(
            f"{model:<8} {summary['n']:>3} "
            f"{summary['exact_accuracy']:>8.3f} {summary['mae']:>7.3f} "
            f"{summary['rmse']:>7.3f} {summary['within_1']:>8.3f} {summary['bias']:>7.3f}"
        )

    print("\nHigh-risk recall (doctor categories 6 ICU, 7 intubation, 8 death):")
    for model, data in report.items():
        recall = data["per_class_recall"]
        bits = [f"{klass}: {recall.get(klass, 0.0):.3f}" for klass in [6, 7, 8]]
        print(f"{model:<8} " + " | ".join(bits))
    return 0


if __name__ == "__main__":
    sys.exit(main())
